#!/usr/bin/env python3
"""Convert Excel data files from the new organized folder structure to data2.js.

Data lives in: Dashboards/data/Data final/ (organized in subfolders)
Output: docs/data2.js
"""
import json
import os
import re
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "Dashboards", "data", "Data final")
OUT_FILE = os.path.join(SCRIPT_DIR, "data.js")
FALLBACK_DATA_JS = os.path.join(SCRIPT_DIR, "data.js")

HEADER_MAP = {
    "Año": "ano",
    "año": "ano",
    "País": "pais",
    "Código ISO": "codigoIso",
    "Código_ISO": "codigoIso",
    "Gini": "gini",
    "Indicador": "indicador",
    "Nivel": "nivel",
    "Valor": "valor",
    "Percentil": "percentil",
    "Participación en el ingreso nacional (%)": "participacionEnElIngresoNacional(%)",
    "Participación en la riqueza nacional (%)": "participacionEnLaRiquezaNacional(%)",
}


def snake_to_camel(name):
    """Convert snake_case to camelCase."""
    parts = name.split("_")
    return parts[0] + "".join(p.capitalize() for p in parts[1:])


def normalize_header(h):
    """Map a raw Excel header to the camelCase key used by app.js."""
    h = str(h).strip()
    if h in HEADER_MAP:
        return HEADER_MAP[h]
    if "_" in h:
        return snake_to_camel(h)
    return h


def read_xlsx(rel_path, sheet=None):
    """Read an xlsx file and return rows as dicts with normalized headers."""
    path = os.path.join(DATA_DIR, rel_path)
    if not os.path.exists(path):
        print(f"  WARNING: missing {rel_path}")
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [normalize_header(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {}
        for h, v in zip(headers, row):
            if v is None:
                record[h] = None
            elif isinstance(v, float):
                record[h] = round(v, 4)
            else:
                record[h] = v
            if h.lower() in ("ano", "anio") and v is not None:
                try:
                    record[h] = int(v)
                except (ValueError, TypeError):
                    pass
        data.append(record)
    return data


def read_multi_sheet(rel_path, sheet_map):
    """Read multiple sheets from one xlsx, return {key: [rows]}."""
    path = os.path.join(DATA_DIR, rel_path)
    if not os.path.exists(path):
        print(f"  WARNING: missing {rel_path}")
        return None
    result = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name, out_key in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not in {rel_path}")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[out_key] = []
            continue
        headers = [normalize_header(h) for h in rows[0]]
        data = []
        for row in rows[1:]:
            record = {}
            for h, v in zip(headers, row):
                if v is None:
                    record[h] = None
                elif isinstance(v, float):
                    record[h] = round(v, 4)
                else:
                    record[h] = v
                if h.lower() in ("ano", "anio") and v is not None:
                    try:
                        record[h] = int(v)
                    except (ValueError, TypeError):
                        pass
            data.append(record)
        result[out_key] = data
    wb.close()
    return result


def dedup(rows):
    """Remove duplicate rows."""
    seen = set()
    unique = []
    for row in rows:
        frozen = json.dumps(row, sort_keys=True, default=str)
        if frozen not in seen:
            seen.add(frozen)
            unique.append(row)
    return unique


def load_fallback_key(key):
    """Extract a dataset from the existing data.js as fallback."""
    if not os.path.exists(FALLBACK_DATA_JS):
        return None
    with open(FALLBACK_DATA_JS, "r", encoding="utf-8") as f:
        content = f.read()
    json_start = content.index("{")
    data = json.loads(content[json_start:-2])
    return data.get(key)


# ── Build all datasets ───────────────────────────────────────────────

print("Reading data from:", DATA_DIR)

datasets = {}

# --- Pobreza ---
datasets["pobrezaTableau"] = read_xlsx("pobreza/Pobreza_tableau.xlsx")
datasets["seriesHistoricas"] = read_xlsx("pobreza/series_historicas_indicadores.xlsx")
datasets["pobrezaProvincial"] = read_xlsx("pobreza/pobreza_provincial.xlsx")
datasets["pobrezaSexoEtnia"] = read_xlsx("pobreza/pobreza_sexo_etnia.xlsx")
datasets["indicadoresSexoEtnia"] = read_xlsx("pobreza/indicadores_sexo_etnia.xlsx")
datasets["pobrezaEducacion"] = read_xlsx("pobreza/pobreza_educacion.xlsx")
datasets["pobrezaEdad"] = read_xlsx("pobreza/pobreza_edad.xlsx")
datasets["pobrezaRegion"] = read_xlsx("pobreza/pobreza_region.xlsx")
datasets["pobrezaMultidimensionalScorecard"] = read_xlsx(
    "pobreza/pobreza_multidimensional_scorecard.xlsx"
)
datasets["pobrezaMultidimensionalSeries"] = read_xlsx(
    "pobreza/pobreza_multidimensional_series.xlsx"
)
datasets["variacionPobrezaSignificancia"] = read_xlsx(
    "pobreza/variacion_pobreza_significancia.xlsx"
)

# --- Ingresos ---
datasets["ingresosSeries"] = read_xlsx("ingresos/ingresos_series.xlsx")
datasets["ingresosArea"] = read_xlsx("ingresos/ingresos_area.xlsx")
datasets["ingresosSexoEtnia"] = read_xlsx("ingresos/ingresos_sexo_etnia.xlsx")
datasets["ingresosEducacion"] = read_xlsx("ingresos/ingresos_educacion.xlsx")
datasets["ingresosEdad"] = read_xlsx("ingresos/ingresos_edad.xlsx")
datasets["ingresosProvincial"] = read_xlsx("ingresos/ingresos_provincial.xlsx")
datasets["ingresosRegion"] = read_xlsx("ingresos/ingresos_region.xlsx")

# --- Ingreso Laboral ---
datasets["inglabSeries"] = read_xlsx("ingresos/inglab_series.xlsx")
datasets["inglabArea"] = read_xlsx("ingresos/inglab_area.xlsx")
datasets["inglabSexoEtnia"] = read_xlsx("ingresos/inglab_sexo_etnia.xlsx")
datasets["inglabEducacion"] = read_xlsx("ingresos/inglab_educacion.xlsx")
datasets["inglabEdad"] = read_xlsx("ingresos/inglab_edad.xlsx")
datasets["inglabProvincial"] = read_xlsx("ingresos/inglab_provincial.xlsx")
datasets["inglabRegion"] = read_xlsx("ingresos/inglab_region.xlsx")

# --- Empleo ---
datasets["empleoSeries"] = read_xlsx("empleo/empleo adecuado/empleo_series.xlsx")
datasets["empleoScorecard"] = read_xlsx("empleo/empleo adecuado/empleo_scorecard.xlsx")
datasets["empleoDemografico"] = read_xlsx(
    "empleo/empleo adecuado/empleo_demografico.xlsx"
)
datasets["variacionEmpleoSignificancia"] = read_xlsx(
    "empleo/empleo adecuado/variacion_empleo_significancia.xlsx"
)

# --- Salarios ---
datasets["salariosSeries"] = read_xlsx("empleo/salarios/salarios_series.xlsx")
datasets["brechasSalariales"] = read_multi_sheet(
    "empleo/salarios/brechas_salariales.xlsx",
    {
        "educacion": "educacion",
        "genero": "genero",
        "genero_civil": "generoCivil",
        "etnia": "etnia",
        "edad": "edad",
    },
)

# --- Distribución del Crecimiento ---
datasets["crecimientoDemografico"] = read_xlsx(
    "distribucion_crecimiento/crecimiento_demografico.xlsx"
)
datasets["crecimientoEmpleoSector"] = read_xlsx(
    "distribucion_crecimiento/crecimiento_empleo_sector.xlsx"
)
datasets["crecimientoPercentiles"] = read_multi_sheet(
    "distribucion_crecimiento/crecimiento_percentiles.xlsx",
    {"percentiles": "percentiles", "deciles": "deciles"},
)
datasets["decilesIngresoAnual"] = read_xlsx(
    "distribucion_crecimiento/deciles_ingreso_anual.xlsx"
)

# --- GIC curves (pre-computed from xlsx folder) ---
def load_gic_curves():
    gic_dir = os.path.join(DATA_DIR, "distribucion_crecimiento", "xlsx")
    if not os.path.isdir(gic_dir):
        print("  WARNING: GIC xlsx folder not found")
        return None
    result = []
    for fname in sorted(os.listdir(gic_dir)):
        if not fname.endswith(".xlsx") or not fname.startswith("gic_"):
            continue
        parts = fname.replace(".xlsx", "").split("_")
        if len(parts) != 4:
            continue
        area = parts[1]  # "nac" or "urb"
        y1, y2 = int(parts[2]), int(parts[3])
        path = os.path.join(gic_dir, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(rows) < 2:
                continue
            for row in rows[1:]:
                if row[0] is None or row[0] == "pctl":
                    continue
                pctl = int(row[0])
                growth = round(float(row[1]), 4) if row[1] is not None else None
                result.append({
                    "area": area,
                    "anioInicio": y1,
                    "anioFin": y2,
                    "pctl": pctl,
                    "growth": growth,
                })
        except Exception as e:
            print(f"  WARNING: error reading {fname}: {e}")
    return result if result else None

datasets["gicCurves"] = load_gic_curves()

# --- Desigualdad: Gini ---
datasets["giniPanel"] = read_xlsx(
    "desigualdad/Gini/gini_panel_tableau.xlsx", "Data"
)
datasets["giniLacComparison"] = read_xlsx("desigualdad/Gini/gini_lac_comparison.xlsx")
datasets["giniTaxImpact"] = read_xlsx("desigualdad/Gini/gini_tax_impact.xlsx")

# --- Desigualdad: Concentración ---
datasets["poblacionPercentiles"] = read_xlsx(
    "desigualdad/Concentracion/poblacion_percentiles.xlsx"
)
datasets["widIngresoPercentiles"] = read_xlsx(
    "desigualdad/Concentracion/WID/WID_ingreso_percentiles_tableau.xlsx"
)
datasets["widRiquezaPercentiles"] = read_xlsx(
    "desigualdad/Concentracion/WID/WID_riqueza_percentiles_tableau.xlsx"
)
datasets["widIngresoPercentilesALC"] = read_xlsx(
    "desigualdad/Concentracion/WID/WID_ingreso_percentiles_ALC_tableau.xlsx"
)
datasets["widRiquezaPercentilesALC"] = read_xlsx(
    "desigualdad/Concentracion/WID/WID_riqueza_percentiles_ALC_tableau.xlsx"
)

# --- Tributación (removed from dashboard) ---
# datasets["tributacionGraficos"] = read_multi_sheet(
#     "tributacion_graficos.xlsx",
#     {
#         "grafico_16_composicion": "grafico16Composicion",
#         "grafico_17_carga": "grafico17Carga",
#     },
# )

# --- Informalidad ---
datasets["informalidadComponentes"] = read_xlsx(
    "empleo/informalidad/Informalidad_y_componentes.xlsx"
)
datasets["informalidadCondiciones"] = read_xlsx(
    "empleo/informalidad/n_condiciones_informalidad.xlsx"
)
datasets["informalidadGenero"] = read_xlsx(
    "empleo/informalidad/Informalidad_genero.xlsx"
)
datasets["informalidadEdad"] = read_xlsx(
    "empleo/informalidad/informalidad_edad.xlsx"
)
datasets["informalidadEtnia"] = read_xlsx(
    "empleo/informalidad/informalidad_etnia.xlsx"
)
datasets["informalidadEducacion"] = read_xlsx(
    "empleo/informalidad/informalidad_educacion.xlsx"
)
datasets["informalidadSexoComponentes"] = read_xlsx(
    "empleo/informalidad/Informalidad_y_componentes_sexo.xlsx"
)

# --- Pobreza Laboral ---
datasets["pobrezaLaboral"] = read_xlsx("empleo/pobreza/pobreza_laboral.xlsx")
datasets["pobrezaLaboralSexo"] = read_xlsx("empleo/pobreza/pobreza_lab_sexo.xlsx")
datasets["pobrezaLaboralArea"] = read_xlsx("empleo/pobreza/pobreza_lab_area.xlsx")
datasets["pobrezaLaboralEduc"] = read_xlsx("empleo/pobreza/pobreza_lab_educ.xlsx")
datasets["pobrezaLaboralEdad"] = read_xlsx("empleo/pobreza/pobreza_lab_edad.xlsx")

# --- SRI Concentración ---
datasets["sriConcentracion"] = read_multi_sheet(
    "desigualdad/Concentracion/SRI/Movilidad y concentracion/Datos graficos todos.xlsx",
    {
        "Concentracion Pretax": "concentracionPretax",
        "Gini": "giniSri",
    },
)
datasets["sriPalma"] = read_multi_sheet(
    "desigualdad/Concentracion/SRI/Movilidad y concentracion/Datos graficos todos.xlsx",
    {"Indice de palma": "indicePalma"},
)

# Capital percentiles from serie_graficos_dina (cleaner format)
datasets["sriCapital"] = read_xlsx(
    "desigualdad/Concentracion/SRI/Movilidad y concentracion/serie_graficos_dina_2010_2024.xlsx"
)

# Transition matrix from dedicated file
datasets["matrizTransicion"] = read_xlsx(
    "desigualdad/Concentracion/SRI/Movilidad y concentracion/Matrices transición/matriz_transicion_deciles_2010_2024_pct.xlsx",
    "Prob_cond_pct",
)

# --- Mortalidad y Violencia ---
datasets["mortalidadJovenes"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_mortalidad_jovenes.xlsx"
)
datasets["mortalidadAdultos"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_mortalidad_adultos.xlsx"
)
datasets["homicidiosJovenes"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_homicidios_jovenes.xlsx"
)
datasets["homicidiosAdultos"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_homicidios_adultos.xlsx"
)
datasets["suicidiosJovenes"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_suicidios_jovenes.xlsx"
)
datasets["suicidiosAdultos"] = read_xlsx(
    "desigualdad/Salud/Mortalidad/jovenes y adultos/por educacion/serie_suicidios_adultos.xlsx"
)

# --- Gini Decomposition ---
datasets["giniDecomposition"] = read_xlsx("desigualdad/Gini/gini_decomposition.xlsx")
datasets["giniDecompositionCuartiles"] = read_xlsx("desigualdad/Gini/gini_decomposition_cuartiles.xlsx")

# --- Palma ALC ---
datasets["palmaAlc"] = read_xlsx("desigualdad/Palma/Palma_ALC_tableau.xlsx", "Data")

# --- Crecimiento Empleo por Sector (all period files) ---
import glob as _glob

_crec_files = _glob.glob(os.path.join(DATA_DIR, "crecimiento_empleo", "datos_graficos_*.xlsx"))
_crec_all = {}
for _fp in sorted(_crec_files):
    _base = os.path.basename(_fp).replace("datos_graficos_", "").replace(".xlsx", "")
    _rows = read_xlsx(os.path.join("crecimiento_empleo", os.path.basename(_fp)), "barras_crecimiento")
    if _rows:
        _crec_all[_base] = _rows
datasets["crecimientoEmpleoPeriodos"] = _crec_all

# --- Informalidad por Rama y Provincia (from LaTeX tables) ---
datasets["informalidadRama"] = [
    {"rama": "Agricultura y pesca", "2001": 96.3, "2005": 97.2, "2010": 94.5, "2015": 88.9, "2020": 94.2, "2025": 94.3},
    {"rama": "Construcción", "2001": 91.4, "2005": 90.8, "2010": 89.2, "2015": 80.8, "2020": 90.1, "2025": 89.2},
    {"rama": "Otros servicios", "2001": 88.6, "2005": 90.7, "2010": 84.2, "2015": 79.8, "2020": 91.5, "2025": 88.9},
    {"rama": "Alojamiento y comidas", "2001": 91.7, "2005": 90.2, "2010": 86.2, "2015": 78.6, "2020": 89.5, "2025": 85.9},
    {"rama": "Hogares como empleadores", "2001": 95.4, "2005": 92.5, "2010": 81.0, "2015": 66.8, "2020": 79.7, "2025": 82.4},
    {"rama": "Transporte y almacenamiento", "2001": 88.8, "2005": 87.2, "2010": 81.5, "2015": 75.8, "2020": 86.8, "2025": 80.6},
    {"rama": "Comercio y rep. vehículos", "2001": 90.4, "2005": 86.9, "2010": 79.5, "2015": 73.0, "2020": 85.1, "2025": 78.7},
    {"rama": "Artes y entretenimiento", "2001": 82.7, "2005": 76.0, "2010": 71.9, "2015": 63.7, "2020": 63.7, "2025": 75.5},
    {"rama": "Activ. profesionales y téc.", "2001": 73.5, "2005": 73.1, "2010": 51.7, "2015": 50.3, "2020": 65.9, "2025": 65.0},
    {"rama": "Manufactura", "2001": 80.2, "2005": 77.8, "2010": 68.3, "2015": 58.0, "2020": 70.9, "2025": 61.2},
    {"rama": "Minas y canteras", "2001": 71.8, "2005": 66.7, "2010": 57.3, "2015": 30.0, "2020": 46.9, "2025": 60.3},
    {"rama": "Servicios administrativos", "2001": 68.1, "2005": 70.0, "2010": 49.2, "2015": 39.4, "2020": 57.7, "2025": 59.5},
    {"rama": "Actividades inmobiliarias", "2001": 8.0, "2005": 5.8, "2010": 0.0, "2015": 45.8, "2020": 38.9, "2025": 53.9},
    {"rama": "Salud y asistencia social", "2001": 54.0, "2005": 47.4, "2010": 40.1, "2015": 25.9, "2020": 39.1, "2025": 36.8},
    {"rama": "Información y comunicaciones", "2001": 64.0, "2005": 64.8, "2010": 55.4, "2015": 42.6, "2020": 42.4, "2025": 29.5},
    {"rama": "Enseñanza", "2001": 50.1, "2005": 40.8, "2010": 27.3, "2015": 14.4, "2020": 16.5, "2025": 22.1},
    {"rama": "Electricidad y gas", "2001": 49.2, "2005": 40.3, "2010": 13.8, "2015": 14.0, "2020": 18.3, "2025": 18.7},
    {"rama": "Actividades financieras", "2001": 39.5, "2005": 25.5, "2010": 24.3, "2015": 13.0, "2020": 15.3, "2025": 14.4},
    {"rama": "Org. extraterritoriales", "2001": 53.6, "2005": 30.4, "2010": 77.0, "2015": 0.0, "2020": 0.0, "2025": 13.7},
    {"rama": "Agua y saneamiento", "2001": 47.0, "2005": 73.7, "2010": 30.1, "2015": 41.2, "2020": 22.1, "2025": 10.1},
    {"rama": "Administración pública", "2001": 29.6, "2005": 16.0, "2010": 12.7, "2015": 4.0, "2020": 4.5, "2025": 3.8},
]

datasets["informalidadProvincia"] = [
    {"provincia": "Pastaza", "2001": 93.9, "2005": 81.5, "2010": 78.0, "2015": 76.9, "2020": 92.7, "2025": 96.6},
    {"provincia": "Napo", "2001": 65.1, "2005": 81.3, "2010": 80.1, "2015": 73.8, "2020": 94.4, "2025": 93.5},
    {"provincia": "Morona Santiago", "2001": 89.6, "2005": 87.5, "2010": 85.1, "2015": 79.4, "2020": 89.8, "2025": 92.0},
    {"provincia": "Orellana", "2001": None, "2005": 91.8, "2010": 84.5, "2015": 75.5, "2020": 80.9, "2025": 91.4},
    {"provincia": "Zamora Chinchipe", "2001": 74.3, "2005": 89.4, "2010": 87.8, "2015": 76.0, "2020": 82.9, "2025": 87.4},
    {"provincia": "Bolívar", "2001": 89.4, "2005": 90.9, "2010": 88.1, "2015": 83.1, "2020": 89.4, "2025": 86.9},
    {"provincia": "Chimborazo", "2001": 94.2, "2005": 88.8, "2010": 86.2, "2015": 84.3, "2020": 91.3, "2025": 86.5},
    {"provincia": "Sucumbíos", "2001": 84.1, "2005": 87.6, "2010": 77.1, "2015": 69.9, "2020": 82.9, "2025": 85.9},
    {"provincia": "Manabí", "2001": 88.6, "2005": 91.6, "2010": 84.2, "2015": 72.8, "2020": 89.1, "2025": 84.5},
    {"provincia": "Esmeraldas", "2001": 88.8, "2005": 87.0, "2010": 79.2, "2015": 70.3, "2020": 84.9, "2025": 84.1},
    {"provincia": "Cotopaxi", "2001": 83.0, "2005": 89.1, "2010": 84.6, "2015": 74.5, "2020": 79.6, "2025": 80.3},
    {"provincia": "Santa Elena", "2001": None, "2005": None, "2010": None, "2015": 73.8, "2020": 87.0, "2025": 80.1},
    {"provincia": "Carchi", "2001": 89.1, "2005": 89.1, "2010": 83.6, "2015": 78.4, "2020": 88.5, "2025": 79.9},
    {"provincia": "Loja", "2001": 91.6, "2005": 85.5, "2010": 81.8, "2015": 72.3, "2020": 81.7, "2025": 79.8},
    {"provincia": "Los Ríos", "2001": 92.4, "2005": 90.8, "2010": 82.2, "2015": 76.8, "2020": 75.0, "2025": 79.7},
    {"provincia": "Cañar", "2001": 86.9, "2005": 87.8, "2010": 83.1, "2015": 71.1, "2020": 78.6, "2025": 79.7},
    {"provincia": "Imbabura", "2001": 85.1, "2005": 83.6, "2010": 78.7, "2015": 70.5, "2020": 87.3, "2025": 79.1},
    {"provincia": "Tungurahua", "2001": 90.8, "2005": 87.4, "2010": 82.0, "2015": 74.1, "2020": 86.2, "2025": 78.5},
    {"provincia": "Santo Domingo de los Tsáchilas", "2001": None, "2005": None, "2010": None, "2015": 73.2, "2020": 83.9, "2025": 76.0},
    {"provincia": "El Oro", "2001": 93.0, "2005": 88.1, "2010": 79.6, "2015": 69.3, "2020": 78.7, "2025": 73.9},
    {"provincia": "Guayas", "2001": 86.0, "2005": 80.2, "2010": 72.3, "2015": 64.3, "2020": 74.1, "2025": 72.4},
    {"provincia": "Azuay", "2001": 83.9, "2005": 79.3, "2010": 70.4, "2015": 66.4, "2020": 70.3, "2025": 69.3},
    {"provincia": "Pichincha", "2001": 72.4, "2005": 69.2, "2010": 60.0, "2015": 49.4, "2020": 65.5, "2025": 60.3},
    {"provincia": "Galápagos", "2001": None, "2005": None, "2010": None, "2015": 42.6, "2020": 84.1, "2025": 45.9},
]

# --- Homicidios Infantiles (parse complex layout) ---
def parse_homicidios_infantiles():
    path = os.path.join(
        DATA_DIR,
        "desigualdad/Salud/Mortalidad/homicidios infantiles/Copia de Datos homicidios.xlsx",
    )
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Tasa homicidios nacional"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return None
    years = [int(v) for v in rows[1][1:13] if v is not None and isinstance(v, (int, float))]
    result = []
    for row in rows[2:]:
        etnia = row[0]
        if etnia is None:
            continue
        vals_total = list(row[1:13])
        vals_nna = list(row[31:43]) if len(row) > 43 else [None] * 12
        for i, y in enumerate(years):
            total = vals_total[i] if i < len(vals_total) else None
            nna = vals_nna[i] if i < len(vals_nna) else None
            if total is not None:
                result.append({"anio": y, "etnia": str(etnia), "tipo": "total", "valor": total})
            if nna is not None:
                result.append({"anio": y, "etnia": str(etnia), "tipo": "nna", "valor": nna})
    return result

datasets["homicidiosInfantiles"] = parse_homicidios_infantiles()

# --- Homicidios NNA (CSV from Boletín 3) ---
import csv

def read_csv_file(rel_path):
    path = os.path.join(SCRIPT_DIR, "..", rel_path)
    if not os.path.exists(path):
        print(f"  WARNING: missing {rel_path}")
        return None
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            record = {}
            for k, v in row.items():
                try:
                    record[k] = round(float(v), 4)
                except (ValueError, TypeError):
                    record[k] = v
            if "anio" in record:
                try:
                    record["anio"] = int(record["anio"])
                except (ValueError, TypeError):
                    pass
            rows.append(record)
    return rows

datasets["homicidiosNna"] = read_csv_file(
    "Boletín 3/3. homicidios_nna/datos_homicidios_nna.csv"
)
datasets["homicidiosJovenesEtnia"] = read_csv_file(
    "Boletín 3/3. homicidios_nna/datos_homicidios_jovenes.csv"
)

# --- Scorecards ---
datasets["scorecards"] = read_xlsx("scorecards_indicadores.xlsx")

# --- Missing files: fall back to existing data.js ---
for key in ["iessAfiliados", "sriPercentilesIngreso"]:
    if key not in datasets or datasets[key] is None:
        print(f"  Falling back to data.js for: {key}")
        datasets[key] = load_fallback_key(key)

# ── Dedup WID datasets ───────────────────────────────────────────────
for key in [
    "widIngresoPercentiles",
    "widRiquezaPercentiles",
    "widIngresoPercentilesALC",
    "widRiquezaPercentilesALC",
]:
    if datasets.get(key):
        datasets[key] = dedup(datasets[key])

# ── Remove None datasets ─────────────────────────────────────────────
datasets = {k: v for k, v in datasets.items() if v is not None}

# ── Write output ─────────────────────────────────────────────────────
with open(OUT_FILE, "w", encoding="utf-8") as f:
    f.write("// Auto-generated from Dashboards/data/Data final/ — do not edit\n")
    f.write("const DATA = ")
    json.dump(datasets, f, ensure_ascii=False, indent=None, default=str)
    f.write(";\n")

print(f"\n✓ Wrote {OUT_FILE}")
for k, v in datasets.items():
    if isinstance(v, list):
        print(f"  {k}: {len(v)} rows")
    elif isinstance(v, dict):
        total = sum(len(sv) for sv in v.values() if isinstance(sv, list))
        print(f"  {k}: dict ({total} total rows)")
