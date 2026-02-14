"""
prepare_tableau_data.py
-----------------------
Transforms raw data files into Tableau-ready long-format tables.
All outputs are saved in the Data/ subfolder with descriptive names.

Usage:
    python3 prepare_tableau_data.py
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

# ──────────────────────────────────────────────────────────────
# PATHS
# ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
DATA = ROOT / "Data"


def log(name: str, df: pd.DataFrame) -> None:
    """Print summary info for a transformed table."""
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"  Shape: {df.shape[0]} rows × {df.shape[1]} cols")
    print(f"  Columns: {list(df.columns)}")
    print(f"  Nulls:\n{df.isnull().sum().to_string()}")
    print(f"  Head:\n{df.head().to_string()}")
    print(f"{'='*60}")


# ──────────────────────────────────────────────────────────────
# 1. gini_panel_tableau.xlsx
#    Union Gini_Nac + Gini_Urb + Gini_rur + gini 90s
# ──────────────────────────────────────────────────────────────
def make_gini_panel() -> pd.DataFrame:
    frames = []

    # Nacional
    df = pd.read_excel(DATA / "Gini_Nac_tableau.xlsx", sheet_name="Data")
    df.columns = ["Año", "pais", "valor"]
    df["nivel"] = "Nacional"
    frames.append(df)

    # Urbano
    df = pd.read_excel(DATA / "Gini_Urb_tableau.xlsx", sheet_name="Data")
    df.columns = ["Año", "pais", "valor"]
    df["nivel"] = "Urbano"
    frames.append(df)

    # Rural
    df = pd.read_excel(DATA / "Gini_rur_tableau.xlsx", sheet_name="Data")
    df.columns = ["Año", "pais", "valor"]
    df["nivel"] = "Rural"
    frames.append(df)

    # 90s urban (from root-level file)
    df90 = pd.read_excel(ROOT / "gini 90s.xlsx", sheet_name="Sheet1", header=None,
                         skiprows=1)
    df90.columns = ["Año", "valor"]
    df90["pais"] = "Ecuador"
    df90["nivel"] = "Urbano"
    df90 = df90[["Año", "pais", "valor", "nivel"]]
    frames.append(df90)

    panel = pd.concat(frames, ignore_index=True)
    # Ensure numeric
    panel["Año"] = pd.to_numeric(panel["Año"], errors="coerce")
    panel["valor"] = pd.to_numeric(panel["valor"], errors="coerce")
    panel.dropna(subset=["Año", "valor"], inplace=True)
    panel["Año"] = panel["Año"].astype(int)
    # Remove duplicates
    panel.drop_duplicates(subset=["Año", "pais", "nivel"], keep="first", inplace=True)

    # Merge nivel + pais into a single 'categoria' column:
    #   ALC → "LAC"
    #   Ecuador + Nacional → "Ecuador"
    #   Ecuador + Urbano  → "Ecuador (Urbano)"
    #   Ecuador + Rural   → "Ecuador (Rural)"
    def _build_categoria(row):
        if row["pais"] == "ALC":
            return "LAC"
        if row["nivel"] == "Nacional":
            return "Ecuador"
        return f"Ecuador ({row['nivel']})"

    panel["categoria"] = panel.apply(_build_categoria, axis=1)
    panel = panel[["Año", "categoria", "valor"]]
    panel.sort_values(["categoria", "Año"], inplace=True)
    panel.reset_index(drop=True, inplace=True)

    return panel


# ──────────────────────────────────────────────────────────────
# 2. gic_deciles_long_tableau.xlsx
#    Pivot period columns → long
# ──────────────────────────────────────────────────────────────
def make_gic_deciles() -> pd.DataFrame:
    df = pd.read_excel(DATA / "Crecimiento_por_deciles_tableau.xlsx",
                       sheet_name="Data")

    # Keep only the main columns (Período + 4 period columns)
    period_cols = [c for c in df.columns if c.startswith("20") or c.startswith("19")]
    df_main = df[["Período"] + period_cols].copy()
    df_main.rename(columns={"Período": "Percentil"}, inplace=True)
    df_main.dropna(subset=["Percentil"], inplace=True)

    melted = df_main.melt(id_vars=["Percentil"],
                          var_name="periodo",
                          value_name="crecimiento")
    melted["crecimiento"] = pd.to_numeric(melted["crecimiento"], errors="coerce")
    melted.dropna(subset=["crecimiento"], inplace=True)
    melted.reset_index(drop=True, inplace=True)

    return melted


# ──────────────────────────────────────────────────────────────
# 3. keynesian_long_tableau.xlsx
#    Extract summary table from Descomposición_Keynesiana
# ──────────────────────────────────────────────────────────────
def make_keynesian() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA / "Descomposición_Keynesiana_tableau.xlsx",
                                data_only=True)
    ws = wb["Data"]

    # Find the summary table — it starts with a row where col A = "Period"
    summary_rows = []
    found = False
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if vals[0] == "Period":
            found = True
            continue
        if found:
            if vals[0] is None or vals[0] == "":
                break
            summary_rows.append({
                "Periodo": str(vals[0]),
                "Delta_PIB": float(vals[1]),
                "Demanda": float(vals[2]),
                "Sustitucion_Importaciones": float(vals[3]),
                "Exportaciones": float(vals[4]),
            })
    wb.close()

    df = pd.DataFrame(summary_rows)

    # Melt components → long
    melted = df.melt(id_vars=["Periodo", "Delta_PIB"],
                     var_name="componente",
                     value_name="valor_pct")
    melted.reset_index(drop=True, inplace=True)

    return melted


# ──────────────────────────────────────────────────────────────
# 4. endi_long_tableau.xlsx
#    Pivot test columns → long, adding group labels
# ──────────────────────────────────────────────────────────────
def make_endi() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA / "ENDI_tableau.xlsx", data_only=True)
    ws = wb["Data"]

    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    # First row is headers — clean newlines from names like "Dice\n(12 a 18 meses)"
    headers = [str(h).replace("\n", " ").strip() if h else "" for h in rows_raw[0]]

    # Parse body: section headers have empty data cols
    records = []
    current_group = ""

    for row in rows_raw[1:]:
        vals = list(row)
        label = str(vals[0]).strip() if vals[0] else ""
        data_vals = vals[1:]

        # Check if this is a section header (all data cols are empty)
        if all(v is None or str(v).strip() == "" for v in data_vals):
            current_group = label
            continue

        for i, hdr in enumerate(headers[1:], start=1):
            v = vals[i]
            if v is not None and str(v).strip() != "":
                records.append({
                    "grupo_poblacion": current_group,
                    "categoria": label,
                    "indicador": hdr,
                    "valor": float(v),
                })

    df = pd.DataFrame(records)
    return df


# ──────────────────────────────────────────────────────────────
# 5–7. Brechas salariales (3 sections)
# ──────────────────────────────────────────────────────────────
def make_brechas() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    wb = openpyxl.load_workbook(DATA / "Brechas_salariales_tableau.xlsx",
                                data_only=True)
    ws = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # ── Section 1: Education (rows 0–27 in 0-indexed) ──
    # Header row 0: Año, Ing_no Univ, Ing_Univ, Salario Básico, ...
    #   Cols K-N (10-13, 0-indexed) = Año_2, Ing_no_Univ_real, Ing_Univ_real, Calificados, Publico, Sexo, Etnia
    # Real wages are in cols 11, 12 (Ing_no Univ_1, Ing_Univ_1 = real values)
    # Gap ratios: col 13 = Calificados, col 14 = Publico, col 15 = Sexo, col 16 = Etnia

    edu_records = []
    for row in all_rows[1:]:  # skip header
        vals = list(row)
        year = vals[10]  # Año_2 column
        if year is None or str(year).strip() == "" or str(year) == "Año":
            # We've hit the next section header
            break
        try:
            year = int(float(str(year)))
        except (ValueError, TypeError):
            continue

        ing_no_univ_real = _to_float(vals[11])
        ing_univ_real = _to_float(vals[12])
        calificados = _to_float(vals[13])
        publico = _to_float(vals[14])
        sexo = _to_float(vals[15])
        etnia = _to_float(vals[16])

        edu_records.append({
            "Año": year,
            "Ing_no_Univ_real": ing_no_univ_real,
            "Ing_Univ_real": ing_univ_real,
            "brecha_calificados": calificados,
            "brecha_publico": publico,
            "brecha_sexo": sexo,
            "brecha_etnia": etnia,
        })

    df_edu = pd.DataFrame(edu_records)

    # ── Section 2: Sex (starts at second "Año" header in nominal cols) ──
    # Rows 29–41 (0-indexed, after "Año, Hombre, Mujer, Total..." header)
    sex_start = None
    for i, row in enumerate(all_rows):
        vals = list(row)
        if i > 5 and vals[0] == "Año" and vals[1] == "Hombre":
            sex_start = i + 1
            break

    sex_records = []
    if sex_start:
        for row in all_rows[sex_start:]:
            vals = list(row)
            year = vals[10]  # Año col in the real-value section
            if year is None or str(year).strip() == "" or str(year) == "Año":
                break
            try:
                year = int(float(str(year)))
            except (ValueError, TypeError):
                continue

            sex_records.append({
                "Año": year,
                "Ing_Hombre_real": _to_float(vals[11]),
                "Ing_Mujer_real": _to_float(vals[12]),
                "brecha_sexo": _to_float(vals[13]),
            })

    df_sex = pd.DataFrame(sex_records)

    # ── Section 3: Ethnicity (starts at "Año, Indígena, No indígena..." header) ──
    eth_start = None
    for i, row in enumerate(all_rows):
        vals = list(row)
        if vals[0] == "Año" and str(vals[1]).startswith("Ind"):
            eth_start = i + 1
            break

    eth_records = []
    if eth_start:
        for row in all_rows[eth_start:]:
            vals = list(row)
            year = vals[10]
            if year is None or str(year).strip() == "":
                break
            try:
                year = int(float(str(year)))
            except (ValueError, TypeError):
                continue

            eth_records.append({
                "Año": year,
                "Ing_Indigena_real": _to_float(vals[11]),
                "Ing_NoIndigena_real": _to_float(vals[12]),
                "brecha_etnia": _to_float(vals[13]),
            })

    df_eth = pd.DataFrame(eth_records)

    return df_edu, df_sex, df_eth


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────────────────────
# 8. impuesto_renta_clean_tableau.xlsx
#    Extract the clean summary section (row 14 onward)
# ──────────────────────────────────────────────────────────────
def make_impuesto_renta() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA / "Impuesto_a_la_renta_tableau.xlsx",
                                data_only=True)
    ws = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the clean summary section starting with "Año", "Límite inferior..."
    start = None
    for i, row in enumerate(all_rows):
        vals = list(row)
        if vals[0] == "Año" and vals[1] is not None and "mite" in str(vals[1]):
            start = i + 1
            break

    if start is None:
        print("WARNING: Could not find the clean tax summary section.")
        return pd.DataFrame()

    records = []
    for row in all_rows[start:]:
        vals = list(row)
        year = vals[0]
        if year is None or str(year).strip() == "":
            break
        try:
            year_num = int(float(str(year)))
        except (ValueError, TypeError):
            continue

        tasa = _to_float(vals[3])
        # Stop if we hit the deflator sub-table (no tasa data)
        if tasa is None:
            break

        records.append({
            "Año": year_num,
            "limite_inf_rango_superior": _to_float(vals[1]),
            "dolares_2015": _to_float(vals[2]),
            "tasa_impositiva": tasa,
            "valor_mensual_rango_sup": _to_float(vals[4]),
        })

    df = pd.DataFrame(records)
    return df


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main() -> None:
    print("=" * 60)
    print("  Preparing Tableau data files")
    print("=" * 60)

    # 1. Gini panel
    gini = make_gini_panel()
    out = DATA / "gini_panel_tableau.xlsx"
    gini.to_excel(out, index=False, sheet_name="Data")
    log("gini_panel_tableau.xlsx", gini)

    # 2. GIC deciles
    gic = make_gic_deciles()
    out = DATA / "gic_deciles_long_tableau.xlsx"
    gic.to_excel(out, index=False, sheet_name="Data")
    log("gic_deciles_long_tableau.xlsx", gic)

    # 3. Keynesian decomposition
    key = make_keynesian()
    out = DATA / "keynesian_long_tableau.xlsx"
    key.to_excel(out, index=False, sheet_name="Data")
    log("keynesian_long_tableau.xlsx", key)

    # 4. ENDI
    endi = make_endi()
    out = DATA / "endi_long_tableau.xlsx"
    endi.to_excel(out, index=False, sheet_name="Data")
    log("endi_long_tableau.xlsx", endi)

    # 5–7. Brechas
    df_edu, df_sex, df_eth = make_brechas()

    out = DATA / "brechas_educacion_tableau.xlsx"
    df_edu.to_excel(out, index=False, sheet_name="Data")
    log("brechas_educacion_tableau.xlsx", df_edu)

    out = DATA / "brechas_sexo_tableau.xlsx"
    df_sex.to_excel(out, index=False, sheet_name="Data")
    log("brechas_sexo_tableau.xlsx", df_sex)

    out = DATA / "brechas_etnia_tableau.xlsx"
    df_eth.to_excel(out, index=False, sheet_name="Data")
    log("brechas_etnia_tableau.xlsx", df_eth)

    # 8. Income tax
    tax = make_impuesto_renta()
    out = DATA / "impuesto_renta_clean_tableau.xlsx"
    tax.to_excel(out, index=False, sheet_name="Data")
    log("impuesto_renta_clean_tableau.xlsx", tax)

    print("\n" + "=" * 60)
    print("  ✅ All files written to Data/")
    print("=" * 60)


if __name__ == "__main__":
    main()
