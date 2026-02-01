import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


def norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def p_value_two_sided(beta: float, se: float) -> float | None:
    if se is None or se == 0:
        return None
    z = beta / se
    return 2.0 * (1.0 - norm_cdf(abs(z)))


def fmt_beta(beta: float, stars: str) -> str:
    return f"{beta:.3f}{stars}"


def fmt_se(se: float) -> str:
    return f"({se:.3f})"


def fmt_p(p: float | None) -> str:
    return "[ ]" if p is None else f"[{p:.3f}]"


def main() -> None:
    # Columns correspond to "Puntaje eliminado": ±0.1, ±0.2, ±0.3, ±0.4, ±0.5
    col_labels = ["(1)", "(2)", "(3)", "(4)", "(5)"]
    puntaje = ["±0.1", "±0.2", "±0.3", "±0.4", "±0.5"]

    # NOTE: These are the "Bias-corrected" coefficients you provided + their SEs.
    rows = [
        ("Óptimo (±6.9123)", [-0.08, -0.146, -0.178, -0.185, -0.23], [0.087, 0.092, 0.094, 0.102, 0.108], ["", "", "*", "*", "**"]),
        ("±1", [0.961, 0.411, -0.017, 1.75, 0.159], [0.251, 0.286, 0.314, 0.495, 0.562], ["***", "", "", "***", ""]),
        ("±2", [0.332, 0.035, -0.112, 0.192, -0.053], [0.163, 0.179, 0.195, 0.254, 0.308], ["**", "", "", "", ""]),
        ("±3", [0.361, 0.21, 0.21, 0.469, 0.439], [0.133, 0.146, 0.155, 0.188, 0.219], ["***", "", "", "**", "**"]),
        ("±4", [0.307, 0.189, 0.185, 0.315, 0.257], [0.116, 0.126, 0.132, 0.153, 0.171], ["***", "", "", "**", ""]),
        ("±5", [0.2, 0.089, 0.062, 0.116, 0.039], [0.103, 0.11, 0.114, 0.128, 0.139], ["*", "", "", "", ""]),
        ("±6", [0.125, 0.022, -0.012, 0.012, -0.061], [0.094, 0.099, 0.102, 0.112, 0.12], ["", "", "", "", ""]),
    ]

    observations = [45888, 45671, 45421, 45184, 44967]
    pseudo_r2 = [".z", ".z", ".z", ".z", ".z"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla 5"

    # Styling
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths (match "table-like" layout)
    ws.column_dimensions["A"].width = 26
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18

    # Title + subtitle
    ws["A1"] = 'Tabla 5. Sensibilidad a las observaciones cercanas al punto de corte (“donut approach”)'
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Óptimo: ±6.912331868086935"
    ws["A2"].font = Font(italic=True)

    # Headers (two-line header like the image)
    header_row_1 = 4
    header_row_2 = 5

    ws.cell(row=header_row_1, column=1, value="Ancho de banda removido").font = Font(bold=True)
    ws.cell(row=header_row_1, column=1).alignment = center

    for j, lab in enumerate(col_labels, start=2):
        c = ws.cell(row=header_row_1, column=j, value=lab)
        c.font = Font(bold=True)
        c.alignment = center

    for j in range(2, 2 + len(col_labels)):
        c = ws.cell(row=header_row_2, column=j, value="dem_index")
        c.alignment = center

    # Body rows
    start_row = 6
    for i, (band_label, betas, ses, stars) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=band_label).alignment = wrap_top
        for j in range(len(col_labels)):
            p = p_value_two_sided(betas[j], ses[j])
            cell_text = "\n".join([fmt_beta(betas[j], stars[j]), fmt_se(ses[j]), fmt_p(p)])
            ws.cell(row=r, column=2 + j, value=cell_text).alignment = wrap_top

    # Footer rows like regression output
    r_obs = start_row + len(rows)
    ws.cell(row=r_obs, column=1, value="Observaciones").alignment = wrap_top
    for j, v in enumerate(observations):
        ws.cell(row=r_obs, column=2 + j, value=str(v)).alignment = center

    r_r2 = r_obs + 1
    ws.cell(row=r_r2, column=1, value="Pseudo R2").alignment = wrap_top
    for j, v in enumerate(pseudo_r2):
        ws.cell(row=r_r2, column=2 + j, value=v).alignment = center

    r_punt = r_r2 + 1
    ws.cell(row=r_punt, column=1, value="Puntaje eliminado").alignment = wrap_top
    for j, v in enumerate(puntaje):
        ws.cell(row=r_punt, column=2 + j, value=v).alignment = center

    # Borders for the whole table area (from header_row_1 to r_punt)
    for r in range(header_row_1, r_punt + 1):
        for c in range(1, 2 + len(col_labels)):
            ws.cell(row=r, column=c).border = border
            if r in (header_row_1, header_row_2):
                ws.cell(row=r, column=c).alignment = center

    # Footnotes
    foot = r_punt + 2
    ws.cell(row=foot, column=1, value="Errores estándar en paréntesis y p-valores en corchetes.")
    ws.cell(row=foot + 1, column=1, value="*** p<.01, ** p<.05, * p<.1")

    # Save next to this script (project root if you run it there)
    out_path = Path("tabla_donut.xlsx").resolve()
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()



