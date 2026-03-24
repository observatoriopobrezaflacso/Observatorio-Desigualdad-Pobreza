#!/usr/bin/env python3
"""
Exporta matrices de transición de deciles a Excel (probabilidades condicionales).

Para cada combinación (año base del .xls, end_year), filtra filas, arma tabla
doble entrada (filas = decil en el año base, columnas = decil de destino). Los valores
en celda son proporciones 0–1 (freq/total_origen), coherentes con Stata; Excel las muestra
como porcentaje con formato 0.00 %.
"""

from __future__ import annotations

import glob
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_GLOB = (
    "/Users/vero/Library/CloudStorage/GoogleDrive-observatorio.pobreza@flacso.edu.ec/"
    "Mi unidad/SRI*/Resultados/20.03.2026/pretax_new/{start}/matriz_transicion_todos_{start}.xls"
)

OUT_DIR = Path(
    "/Users/vero/Library/CloudStorage/GoogleDrive-observatorio.pobreza@flacso.edu.ec/"
    "Mi unidad/Boletín 2/Resultados"
)

# (año base del archivo / columna decil_origen, end_year destino)
JOBS: list[tuple[int, int]] = [
    (2010, 2014),
    (2010, 2017),
    (2010, 2024),
    (2014, 2024),
    (2017, 2024),
]

SHEET_NAME = "Prob_cond_pct"

_FILL_HEADER = PatternFill("solid", fgColor="FF2C3E50")
_FILL_ROW_ALT = PatternFill("solid", fgColor="FFF8F9FA")
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
_FONT_CORNER = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
_FONT_BODY_LABEL = Font(name="Calibri", size=10, bold=True)
_FONT_BODY_NUM = Font(name="Calibri", size=10)
_THIN = Side(style="thin", color="FFBDC3C7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
# Excel: formato porcentaje estándar (el valor en celda debe ser proporción 0–1, p. ej. 0.0994 → 9.94 %).
_NUM_FMT_PCT = "0.00%"


def _source_path(start_year: int) -> str:
    pattern = BASE_GLOB.format(start=start_year)
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No se encontró archivo: {pattern}")
    return matches[0]


def _decil_origen_col(start_year: int) -> str:
    return f"decil_{start_year}"


def _format_matrix_sheet(ws, n_data_rows: int, n_data_cols: int) -> None:
    max_row = 1 + n_data_rows
    max_col = n_data_cols

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 10.0 if c > 1 else 22.0

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER

            if r == 1:
                cell.fill = _FILL_HEADER
                cell.font = _FONT_HEADER if c > 1 else _FONT_CORNER
                cell.alignment = _ALIGN_CENTER
                continue

            if c == 1:
                cell.font = _FONT_BODY_LABEL
                cell.alignment = _ALIGN_LEFT
                if r % 2 == 0:
                    cell.fill = _FILL_ROW_ALT
            else:
                cell.font = _FONT_BODY_NUM
                cell.number_format = _NUM_FMT_PCT
                cell.alignment = _ALIGN_CENTER
                if r % 2 == 0:
                    cell.fill = _FILL_ROW_ALT

    ws.row_dimensions[1].height = 28


def export_matrix(start_year: int, end_year: int) -> Path:
    path = _source_path(start_year)
    origen = _decil_origen_col(start_year)

    df = pd.read_excel(path, sheet_name=0, header=0)
    ey = pd.to_numeric(df["end_year"], errors="coerce")
    t1 = pd.to_numeric(df["anio_t1_v"], errors="coerce")
    ini = pd.to_numeric(df["ini_year"], errors="coerce")
    df = df[(ey == end_year) & (t1 == end_year) & (ini == start_year)].copy()
    if len(df) != 100:
        raise ValueError(
            f"{path}: se esperaban 100 filas (10×10) con end_year==anio_t1_v=={end_year} "
            f"e ini_year=={start_year}; hay {len(df)}."
        )

    df[origen] = pd.to_numeric(df[origen], errors="coerce").astype(int)
    df["decil_dest"] = pd.to_numeric(df["decil_dest"], errors="coerce").astype(int)
    freq = pd.to_numeric(df["freq"], errors="coerce")
    tot = pd.to_numeric(df["total_origen"], errors="coerce")
    prob_file = pd.to_numeric(df["prob_cond"], errors="coerce")
    prob = freq / tot
    diff = (prob - prob_file).abs().max()
    if diff > 1e-6:
        raise ValueError(
            f"{path}: prob_cond del archivo no coincide con freq/total_origen (max |Δ|={diff})."
        )

    wide = df.assign(_prob=prob).pivot(index=origen, columns="decil_dest", values="_prob")
    wide = wide.reindex(index=range(1, 11), columns=range(1, 11))

    corner = f"Decil {start_year}/Decil {end_year}"
    out = wide.reset_index()
    out = out.rename(
        columns={
            origen: corner,
            **{c: str(c) for c in range(1, 11)},
        }
    )

    out_path = OUT_DIR / f"matriz_transicion_deciles_{start_year}_{end_year}_pct.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        ws = writer.sheets[SHEET_NAME]
        _format_matrix_sheet(ws, n_data_rows=len(out), n_data_cols=out.shape[1])

    return out_path


def main() -> None:
    for start_year, end_year in JOBS:
        out = export_matrix(start_year, end_year)
        print(f"Escrito: {out}")


if __name__ == "__main__":
    main()
